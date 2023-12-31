"""
 This is the default license template.
 
 File: scripts_generator.py
 Author: ashishsingh
 Copyright (c) 2023 ashishsingh
 
 To edit this license information: Press Ctrl+Shift+P and press 'Create new License Template...'.
"""

# MultiLingual xml generator from Excel
# pip install openpyxl --->  install this module

import os
import sys
from importlib import reload
reload(sys)

import openpyxl 
wb = openpyxl.load_workbook('Multilingual.xlsx')
sheet = wb.get_sheet_by_name('Consumer')
totalRows = 130

def strformator(s,s1):
    a=s.split(" ")
    final_str = ""
    j=1
    for i in s1:
        if i == '*':
            final_str += a[j-1]
            j+=1
        elif i == "{":
            final_str += '"<>'
        elif i == "}":
            final_str += '<>"'
        else:
            final_str+=i
    return final_str

        


blacklist = ['\n','"']

f2=open("../src/Multilingual/Types.purs", 'w+')
print("module Language.Types where\n\nimport Prelude\n\nimport Control.Monad.List.Trans (Step)",end="\n\n", file=f2)
print("data STR = ",end="", file=f2)
for k in range(2 ,4):
    lang = sheet.cell(row=2, column=k).value
    print("Language => " + lang)
    f1=open("../src/Multilingual/Language/"+lang+".purs", 'w+')

    print("module "+ lang +" where\n\nimport Prelude\n\nimport Language.Types (STR(..))",end="\n\n", file=f1)
    print("get"+lang+" ::STR -> String\nget"+lang +" script = case script of",end="\n", file=f1)
    for l in range(3, totalRows + 1):
        key = sheet.cell(row=l, column=1).value
        if key != None:
            if k==2:
                # print(key,end="\n", file=f2)
                if l<totalRows:
                    temp = ' '.join(key.splitlines()[0:])
                temp1 = temp.split(' ')
                if (len(temp1) > 1):
                    for i in range(0,len(temp1)):
                        if(i == 0):
                            print(""+temp1[0]+"",end= " " , file=f2)
                        elif (i == len(temp1)-1):

                            print("String",end= " " , file=f2)
                        else:
                            print("String",end= " " , file=f2)

                else :
                    print(key,end="", file=f2)
                if l<totalRows :
                    print("\n\t\t | ",end="", file=f2)
                    
            val = sheet.cell(row=l, column=k).value
            if val==None:
                val = " "
            strKey = ' '.join(key.splitlines()[0:])
            strkey1 = ' '.join(str(val).splitlines()[0:])
            aa = strkey1[:(len(strkey1))]
            key3 = strformator(strKey,aa)
            print("\t\t"+key+" -> \""+key3+"\"",end="\n", file=f1)
            print("",end="",file=f2)    
    print("\n",end="\n",file=f2)   
    for l in range(3, totalRows+1):
        key = sheet.cell(row=l, column=1).value
        if k==2 and key != None:
            print("\ntype "+key+" = String",end="\n",file=f2)         